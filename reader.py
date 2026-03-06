"""
UpdateAlarm - Excel okuyucu (gelismis veri modeli)
update_alarm.py mantigi korunarak genisletildi.
"""
import datetime as dt
import re
import openpyxl
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

# ──────────────────────────── Turkce yardimci ────────────────────────────────
TR_MONTHS = {
    "ocak": 1, "subat": 2, "mart": 3, "nisan": 4,
    "mayis": 5, "haziran": 6, "temmuz": 7, "agustos": 8,
    "eylul": 9, "ekim": 10, "kasim": 11, "aralik": 12,
    # unicode varyantlari
    "şubat": 2, "mayıs": 5, "ağustos": 8, "eylül": 9,
    "kasım": 11, "aralık": 12,
}
TR_WEEKDAYS = {
    "pazartesi": 0,
    "sali": 1, "salı": 1,
    "carsamba": 2, "çarşamba": 2,
    "persembe": 3, "perşembe": 3,
    "cuma": 4,
    "cumartesi": 5,
    "pazar": 6,
}
TR_WEEKDAY_NAMES = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma", "Cumartesi", "Pazar"]


def normalize(s: str) -> str:
    s = (s or "").strip().lower()
    # unicode normalize
    for a, b in [("ş","s"),("ı","i"),("ğ","g"),("ü","u"),("ö","o"),("ç","c"),("İ","i"),("Ş","s"),("Ğ","g"),("Ü","u"),("Ö","o"),("Ç","c")]:
        s = s.replace(a, b)
    return " ".join(s.split())


# ──────────────────────────── Veri modelleri ─────────────────────────────────
@dataclass
class UpdateEvent:
    cycle: str
    cycle_num: Optional[int]
    rule: str
    window: str
    when: dt.datetime
    server: str
    pre_mails: List[str] = field(default_factory=list)
    post_mails: List[str] = field(default_factory=list)

    @property
    def urgency(self) -> str:
        delta = (self.when.date() - dt.date.today()).days
        if delta < 0:   return "past"
        if delta == 0:  return "today"
        if delta <= 2:  return "soon"
        if delta <= 7:  return "week"
        if delta <= 30: return "month"
        return "future"

    @property
    def time_remaining(self) -> str:
        delta = self.when - dt.datetime.now()
        if delta.total_seconds() < 0:
            return "Gecti"
        d, rem = divmod(int(delta.total_seconds()), 86400)
        h, rem = divmod(rem, 3600)
        m = rem // 60
        if d > 0:   return f"{d}g {h}s"
        if h > 0:   return f"{h}s {m}dk"
        return f"{m}dk"

    @property
    def day_tr(self) -> str:
        return TR_WEEKDAY_NAMES[self.when.weekday()]


@dataclass
class CycleGroup:
    name: str
    num: Optional[int]
    servers: List[str]
    rules: List[str]
    pre_mails: List[str]
    post_mails: List[str]
    events: List[UpdateEvent]

    @property
    def next_event(self) -> Optional[UpdateEvent]:
        now = dt.datetime.now()
        fut = [e for e in self.events if e.when >= now]
        return min(fut, key=lambda e: e.when) if fut else None

    @property
    def server_count(self) -> int:
        return len(set(self.servers))


@dataclass
class AppData:
    events: List[UpdateEvent]
    cycles: List[CycleGroup]
    load_time: dt.datetime
    source_file: str
    error: Optional[str] = None

    @property
    def upcoming_today(self) -> List[UpdateEvent]:
        today = dt.date.today()
        return [e for e in self.events if e.when.date() == today and e.when >= dt.datetime.now()]

    @property
    def upcoming_week(self) -> List[UpdateEvent]:
        now = dt.datetime.now()
        cutoff = now + dt.timedelta(days=7)
        return [e for e in self.events if now <= e.when <= cutoff]

    @property
    def upcoming_month(self) -> List[UpdateEvent]:
        now = dt.datetime.now()
        cutoff = now + dt.timedelta(days=30)
        return [e for e in self.events if now <= e.when <= cutoff]


# ──────────────────────────── Tarih parse yardimcilari ───────────────────────
def _to_time(h: str, m: str) -> dt.time:
    hh, mm = int(h), int(m)
    if hh == 24 and mm == 0: hh = 0
    if not (0 <= hh <= 23 and 0 <= mm <= 59):
        return dt.time(9, 0)
    return dt.time(hh, mm)


def parse_time_window(text) -> Tuple[dt.time, str]:
    # openpyxl bazen datetime.time nesnesi döndürür
    if isinstance(text, dt.time):
        t = text
        return t, t.strftime("%H:%M")
    matches = re.findall(r"(\d{1,2})\s*[:.]\s*(\d{2})", str(text))
    if not matches:
        t = dt.time(9, 0)
        return t, t.strftime("%H:%M")
    start = _to_time(matches[0][0], matches[0][1])
    window = start.strftime("%H:%M")
    if len(matches) >= 2:
        end = _to_time(matches[1][0], matches[1][1])
        window = f"{start:%H:%M}-{end:%H:%M}"
    return start, window


def parse_explicit_dates(rule_text: str) -> List[dt.date]:
    out: List[dt.date] = []
    for d, mon, y in re.findall(
        r"(\d{1,2})\s+([A-Za-z\u00c0-\u024f]+)\s+(\d{4})", rule_text or "", flags=re.IGNORECASE
    ):
        month = TR_MONTHS.get(normalize(mon))
        if month:
            try: out.append(dt.date(int(y), month, int(d)))
            except ValueError: pass
    return sorted(set(out))


def nth_weekday_of_month(year: int, month: int, weekday: int, n: int) -> Optional[dt.date]:
    first = dt.date(year, month, 1)
    shift = (weekday - first.weekday()) % 7
    day = 1 + shift + 7 * (n - 1)
    try: return dt.date(year, month, day)
    except ValueError: return None


def last_weekday_of_month(year: int, month: int, weekday: int) -> dt.date:
    if month == 12: last = dt.date(year, 12, 31)
    else: last = dt.date(year, month + 1, 1) - dt.timedelta(days=1)
    return last - dt.timedelta(days=(last.weekday() - weekday) % 7)


def month_iter(start: dt.date):
    y, m = start.year, start.month
    while True:
        yield y, m
        m += 1
        if m == 13: m = 1; y += 1


def _find_weekday(txt: str) -> Optional[int]:
    for k, v in TR_WEEKDAYS.items():
        if k in txt: return v
    return None


def _second_tuesday(year: int, month: int) -> Optional[dt.date]:
    return nth_weekday_of_month(year, month, TR_WEEKDAYS["sali"], 2)


def _weekday_after_second_tuesday(year: int, month: int, target_wd: int) -> Optional[dt.date]:
    base = _second_tuesday(year, month)
    if not base: return None
    delta = (target_wd - base.weekday()) % 7
    if delta == 0: delta = 7
    return base + dt.timedelta(days=delta)


def next_occurrences(rule_text: str, base_date: dt.date, count: int = 120) -> List[dt.date]:
    raw = rule_text or ""
    txt = normalize(raw)
    explicit = parse_explicit_dates(raw)
    if explicit:
        return [d for d in explicit if d >= base_date][:count]

    out: List[dt.date] = []

    # Ayın N. günü
    m = re.search(r"\bayin\s+(\d{1,2})\.\s*gunu\b", txt)
    if m:
        target = int(m.group(1))
        for y, mo in month_iter(base_date):
            try:
                d = dt.date(y, mo, target)
                if d >= base_date: out.append(d)
            except ValueError: pass
            if len(out) >= count: break
        return out

    # Son weekday
    if "son" in txt:
        wd = _find_weekday(txt)
        if wd is not None and re.search(r"\bson\s+(pazartesi|sali|carsamba|persembe|cuma|cumartesi|pazar)\b", txt):
            for y, mo in month_iter(base_date):
                d = last_weekday_of_month(y, mo, wd)
                if d >= base_date: out.append(d)
                if len(out) >= count: break
            return out

    # Ilk hafta weekday
    if "ayin ilk haftasi" in txt or "ayın ilk haftası" in txt:
        wd = _find_weekday(txt)
        if wd is None: wd = TR_WEEKDAYS["cuma"]
        for y, mo in month_iter(base_date):
            d = nth_weekday_of_month(y, mo, wd, 1)
            if d and d >= base_date: out.append(d)
            if len(out) >= count: break
        return out

    # 2. Salidan N gun sonra
    m_after = re.search(r"\b2\.\s*sali[a-z]*dan\s+(\d{1,2})\s*gun\s+sonra", txt)
    if m_after:
        days_after = int(m_after.group(1))
        for y, mo in month_iter(base_date):
            base = _second_tuesday(y, mo)
            if base:
                d = base + dt.timedelta(days=days_after)
                if d >= base_date: out.append(d)
            if len(out) >= count: break
        return out

    # 2. Saliyi takip eden weekday
    if ("2.sali" in txt or "2. sali" in txt) and "takip" in txt:
        target_wd = None
        for k, v in TR_WEEKDAYS.items():
            if k in txt and k not in ("sali", "salı"):
                target_wd = v; break
        if target_wd is None: target_wd = TR_WEEKDAYS["cuma"]
        for y, mo in month_iter(base_date):
            d = _weekday_after_second_tuesday(y, mo, target_wd)
            if d and d >= base_date: out.append(d)
            if len(out) >= count: break
        return out

    # 2. Salinin Cuma aksamı (varyant)
    if ("2. salisinin" in txt or "2.sali" in txt) and "cuma" in txt and ("aksam" in txt or "akşam" in txt):
        for y, mo in month_iter(base_date):
            d = _weekday_after_second_tuesday(y, mo, TR_WEEKDAYS["cuma"])
            if d and d >= base_date: out.append(d)
            if len(out) >= count: break
        return out

    # 2. Salinin haftasonu
    if ("2. sali" in txt or "2.sali" in txt) and ("haftasonu" in txt or "hafta sonu" in txt):
        for y, mo in month_iter(base_date):
            d = _weekday_after_second_tuesday(y, mo, TR_WEEKDAYS["cumartesi"])
            if d and d >= base_date: out.append(d)
            if len(out) >= count: break
        return out

    # Genel: Ayin N. weekday
    m_nth = re.search(r"\bayin\s+([1-5])\.\s*([a-z]+)\b", txt)
    if m_nth:
        n = int(m_nth.group(1))
        wd = _find_weekday(txt)
        if wd is not None:
            for y, mo in month_iter(base_date):
                d = nth_weekday_of_month(y, mo, wd, n)
                if d and d >= base_date: out.append(d)
                if len(out) >= count: break
            return out

    return out


# ──────────────────────────── Excel okuyucu ──────────────────────────────────
def _cycle_num(name: str) -> Optional[int]:
    m = re.search(r"(\d+)\s*$", name or "")
    return int(m.group(1)) if m else None


def _extract_cycle_name(row_text: str) -> Optional[str]:
    m = re.search(r"(.{0,120}?update\s*cycle\s*\d+)", row_text, flags=re.IGNORECASE)
    if not m: return None
    name = re.sub(r"\s*hk\.?\s*$", "", m.group(1).strip(), flags=re.IGNORECASE).strip()
    return " ".join(name.split())


def _server_name(cell_text: str) -> str:
    s = str(cell_text or "").strip()
    # Parantez ve sonrasını sil: "EUGBZPBIPRX3(THEOBAOLD OLMUŞ...)" → "EUGBZPBIPRX3"
    s = re.sub(r'\s*\(.*', '', s).strip()
    # "10.237.75.16 - EUGBZISSQL" → "EUGBZISSQL"
    if " - " in s:
        s = s.split(" - ", 1)[1].strip()
    # "10.237.74.137--EUGBZNESSUS" → "EUGBZNESSUS"
    elif re.search(r'-{2,}', s):
        s = re.split(r'-{2,}', s, 1)[-1].strip()
    # "10.237.76.45 EUGBZPICUSNEXT" (IP + boşluk) → "EUGBZPICUSNEXT"
    elif re.match(r'^\d+\.\d+\.\d+\.\d+\s+', s):
        s = re.sub(r'^\d+\.\d+\.\d+\.\d+\s+', '', s).strip()
    # Sadece kontrol karakterlerini temizle
    s = re.sub(r"[\x00-\x1f\x7f]", "", s).strip()
    return s


def _split_mails(text: str) -> List[str]:
    raw = str(text or "").strip()
    if not raw: return []
    seen, out = set(), []
    for p in re.split(r"[,\n;\s\u00a0]+", raw):
        p = p.strip().strip("\u00a0")
        if p and "@" in p and p.lower() not in seen:
            seen.add(p.lower()); out.append(p)
    return out


def load_excel(path: str, horizon_days: int = 365) -> AppData:
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return AppData(events=[], cycles=[], load_time=dt.datetime.now(), source_file=path, error=str(e))

    ws = wb[wb.sheetnames[0]]
    now = dt.datetime.now()
    today = now.date()
    horizon = today + dt.timedelta(days=horizon_days)

    events: List[UpdateEvent] = []
    cycle_events: Dict[str, List[UpdateEvent]] = {}
    cycle_servers: Dict[str, List[str]] = {}
    cycle_rules: Dict[str, List[str]] = {}
    cycle_pre: Dict[str, List[str]] = {}
    cycle_post: Dict[str, List[str]] = {}

    current_cycle = "Update Cycle ?"
    current_rule: Optional[str] = None
    current_time: Optional[str] = None

    def _add_mails(cycle, pre_txt, post_txt):
        cycle_pre.setdefault(cycle, [])
        cycle_post.setdefault(cycle, [])
        for e in _split_mails(pre_txt):
            if e.lower() not in {x.lower() for x in cycle_pre[cycle]}:
                cycle_pre[cycle].append(e)
        for e in _split_mails(post_txt):
            if e.lower() not in {x.lower() for x in cycle_post[cycle]}:
                cycle_post[cycle].append(e)

    for r in range(1, ws.max_row + 1):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value

        row_text = " ".join(str(ws.cell(r, col).value or "") for col in range(1, 13))
        cyc = _extract_cycle_name(row_text)
        if cyc:
            current_cycle = cyc

        _add_mails(current_cycle, c4, c5)

        if isinstance(c1, str) and normalize(c1) == "update gecilecek sunucular":
            current_rule = None; current_time = None; continue

        # Tamamen bos satir: sadece gec, kurali sifirLAMA
        # (bos satir Excel'de görsel ayirici olarak kullanilabiliyor;
        #  sifirlamak sonraki sunuculari kaybettirir)
        if all(v is None for v in (c1, c2, c3, c4, c5)):
            continue

        # c1 bos ya da None ise bu satirda sunucu yok
        if c1 is None or not str(c1).strip():
            continue

        server = _server_name(str(c1))
        if not server:
            continue

        if c2 is not None and str(c2).strip():
            current_rule = str(c2).strip()
        if c3 is not None:
            # datetime.time nesnesini doğrudan sakla, string ise strip et
            if isinstance(c3, dt.time):
                current_time = c3
            elif str(c3).strip():
                current_time = str(c3).strip()

        # Sunucuyu her halukarda cycle listesine ekle (event üretilip üretilmemesinden bağımsız)
        if server not in cycle_servers.get(current_cycle, []):
            cycle_servers.setdefault(current_cycle, []).append(server)

        if not current_rule or not current_time:
            # Kural/saat bilinmiyor ama sunucu zaten cycle'a eklendi
            continue

        if current_rule not in cycle_rules.get(current_cycle, []):
            cycle_rules.setdefault(current_cycle, []).append(current_rule)

        start_t, window = parse_time_window(current_time)

        if isinstance(c2, dt.datetime):
            dates = [c2.date()]
        elif isinstance(c2, dt.date):
            dates = [c2]
        else:
            dates = next_occurrences(current_rule, today, count=120)

        for d in dates:
            if d < today or d > horizon:
                continue
            when = dt.datetime.combine(d, start_t)
            if when < now:
                continue

            pre = cycle_pre.get(current_cycle, [])
            post = cycle_post.get(current_cycle, [])

            ev = UpdateEvent(
                cycle=current_cycle,
                cycle_num=_cycle_num(current_cycle),
                rule=current_rule,
                window=window,
                when=when,
                server=server,
                pre_mails=list(pre),
                post_mails=list(post),
            )
            events.append(ev)
            cycle_events.setdefault(current_cycle, []).append(ev)

    events.sort(key=lambda e: (e.when, e.cycle, e.server))

    cycles: List[CycleGroup] = []
    for cname in sorted(cycle_events.keys(), key=lambda x: (_cycle_num(x) or 999, x)):
        cycles.append(CycleGroup(
            name=cname,
            num=_cycle_num(cname),
            servers=sorted(set(cycle_servers.get(cname, []))),
            rules=cycle_rules.get(cname, []),
            pre_mails=cycle_pre.get(cname, []),
            post_mails=cycle_post.get(cname, []),
            events=cycle_events.get(cname, []),
        ))

    return AppData(events=events, cycles=cycles, load_time=now, source_file=path)
